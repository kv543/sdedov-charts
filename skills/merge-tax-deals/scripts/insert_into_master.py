#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
insert_into_master.py — מכניס שורות מקובץ העסקאות המאוחד (פלט של merge_tax_files.py)
אל קובץ המכירות הכללי, לגיליון "בסיס נתוני מכירות", עם כל לוגיקת העיבוד:
  • מיפוי 10 עמודות רשות-המיסים -> 21 עמודות המאסטר (כולל נוסחאות).
  • סיווג "סוג עסקה": מימוש אופציה אם מרכז + יש אופציה תואמת (גוש-חלקה מלא) ביוני 2025–מרץ 2026; אחרת דירה.
  • גילום מכירות חלק יחסי (חלק נמכר < 1): תמורה ושווי × round(1/חלק), חלק נמכר = 1.
  • מחיקת כפילויות מדויקות + איחוד שורות חלק-יחסי של אותה דירה.
  • מיון כל הגיליון לפי תאריך, וחילול מחדש של כל הנוסחאות מתבניות (שומר $U$2:$U$5).

שימוש:
  python3 insert_into_master.py <master.xlsx> <merged_tax.xlsx> [output.xlsx]
"""
import sys, datetime, re
import openpyxl

SHEET = 'בסיס נתוני מכירות'
ESHKOL_GUSHIM = {'6634', '7186'}           # מתחם אשכול
OPT_FROM = (2025, 6)                        # חלון אופציות: יוני 2025
OPT_TO   = (2026, 3)                        # עד מרץ 2026

# מיפוי כותרת-מאסטר -> אינדקס עמודה (1-based), נקבע דינמית מהכותרת
MASTER_INPUT_FROM_TAX = {   # כותרת במאסטר : תת-מחרוזת לזיהוי בקובץ המאוחד
    'גוש חלקה':        'גוש חלקה',
    'יום מכירה':       'יום מכירה',
    'תמורה מוצהרת':    'תמורה מוצהרת',
    'שווי מכירה':      'שווי מכירה',
    'מהות':            'מהות',
    'חלק נמכר':        'חלק נמכר',
    'ישוב':            'ישוב',
    'שנת בניה':        'שנת בניה',
    'שטח':             'שטח',
    'חדרים':           'חדרים',
}

NUMFMT_ACCT0  = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
NUMFMT_ACCT2  = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
NUMFMT_DATE   = 'mm-dd-yy'


def col_finder(headers):
    def find(sub, excl=None):
        for i, h in enumerate(headers):
            if h and sub in str(h) and (excl is None or excl not in str(h)):
                return i + 1
        raise KeyError(sub)
    return find


def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, datetime.date):
        return datetime.datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        s = v.strip()
        for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d'):
            try:
                return datetime.datetime.strptime(s, fmt)
            except ValueError:
                pass
    return None


def gush_of(gush_chelka):
    """'006634-0015-263-00' -> '6634'"""
    first = str(gush_chelka).split('-')[0]
    try:
        return str(int(first))
    except ValueError:
        return first.lstrip('0')


def compound_of(gush_chelka):
    return 'אשכול' if gush_of(gush_chelka) in ESHKOL_GUSHIM else 'מרכז'


def in_option_window(dt):
    if dt is None:
        return False
    ym = (dt.year, dt.month)
    return OPT_FROM <= ym <= OPT_TO


# ----------------------------------------------------------------------------- read tax
def read_tax_rows(merged_path):
    wb = openpyxl.load_workbook(merged_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    find = col_finder(headers)
    idx = {k: find(sub) for k, sub in [
        ('a', 'גוש חלקה'), ('f', 'יום מכירה'), ('j', 'תמורה מוצהרת'),
        ('k', 'שווי מכירה'), ('l', 'מהות'), ('n', 'חלק נמכר'),
        ('o', 'ישוב'), ('p', 'שנת בניה'), ('q', 'שטח'), ('r', 'חדרים')]}
    rows = []
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, idx['a']).value
        if a in (None, ''):
            continue
        rows.append({
            'a': str(a).strip(),
            'f': parse_date(ws.cell(r, idx['f']).value),
            'fstr': (parse_date(ws.cell(r, idx['f']).value).strftime('%d/%m/%Y')
                     if parse_date(ws.cell(r, idx['f']).value) else None),
            'j': ws.cell(r, idx['j']).value,
            'k': ws.cell(r, idx['k']).value,
            'l': ws.cell(r, idx['l']).value,
            'n': ws.cell(r, idx['n']).value,
            'o': ws.cell(r, idx['o']).value,
            'p': ws.cell(r, idx['p']).value,
            'q': ws.cell(r, idx['q']).value,
            'r': ws.cell(r, idx['r']).value,
        })
    return rows


# ----------------------------------------------------------------- clean / dedupe / gross-up
def clean_tax_rows(rows, log):
    # 1) מחיקת כפילויות מדויקות (כל השדות זהים)
    seen = {}
    step1 = []
    removed_exact = 0
    for row in rows:
        key = (row['a'], row['fstr'], row['j'], row['k'], row['n'], row['q'])
        if key in seen:
            removed_exact += 1
            continue
        seen[key] = True
        step1.append(row)

    # 2) איחוד שורות חלק-יחסי של אותה דירה (אותו גוש-חלקה+תאריך, כל השורות חלק<1) -> שורה אחת
    by_unit = {}
    for row in step1:
        by_unit.setdefault((row['a'], row['fstr']), []).append(row)
    step2 = []
    collapsed_frac = 0
    for (a, f), grp in by_unit.items():
        frac = [g for g in grp if isinstance(g['n'], (int, float)) and g['n'] < 1]
        if len(grp) > 1 and len(frac) == len(grp):
            # כל השורות חלק-יחסי -> אותה דירה, שמור אחת (תמורה גבוהה ביותר, דטרמיניסטי)
            keep = max(grp, key=lambda g: (g['j'] or 0))
            step2.append(keep)
            collapsed_frac += len(grp) - 1
        else:
            step2.extend(grp)

    # 3) גילום חלק-יחסי -> דירה שלמה
    grossed = 0
    for row in step2:
        n = row['n']
        if isinstance(n, (int, float)) and 0 < n < 1:
            factor = round(1 / n)            # 0.333 -> 3, 0.5 -> 2, 0.25 -> 4
            if factor >= 2:
                row['j'] = round((row['j'] or 0) * factor)
                row['k'] = round((row['k'] or 0) * factor)
                row['n'] = 1
                grossed += 1

    log.append(f"  כפילויות מדויקות שהוסרו: {removed_exact}")
    log.append(f"  שורות חלק-יחסי שאוחדו: {collapsed_frac}")
    log.append(f"  שורות חלק-יחסי שגולמו ×factor: {grossed}")
    return step2


# --------------------------------------------------------------------------- read master
def read_master(master_path):
    wb = openpyxl.load_workbook(master_path)             # keep formulas/styles
    wbv = openpyxl.load_workbook(master_path, data_only=True)
    ws, wsv = wb[SHEET], wbv[SHEET]
    headers = [c.value for c in ws[1]]
    find = col_finder(headers)
    C = {name: find(*spec) if isinstance(spec, tuple) else find(spec) for name, spec in {
        'A': 'גוש חלקה', 'E': 'מתחם', 'F': 'יום מכירה',
        'J': ('תמורה מוצהרת', 'טווח'), 'K': 'שווי מכירה', 'L': 'מהות',
        'M': 'סוג עסקה', 'N': 'חלק נמכר', 'O': 'ישוב', 'P': 'שנת בניה',
        'Q': 'שטח', 'R': 'חדרים', 'U': None}.items() if name != 'U'}
    # column U = the first column with no header (the thresholds table)
    U = None
    for i, h in enumerate(headers):
        if h is None:
            U = i + 1
            break
    existing = []
    for r in range(2, ws.max_row + 1):
        a = wsv.cell(r, C['A']).value
        if a in (None, ''):
            continue
        existing.append({
            'a': str(a).strip(),
            'f': parse_date(wsv.cell(r, C['F']).value),
            'j': wsv.cell(r, C['J']).value, 'k': wsv.cell(r, C['K']).value,
            'l': wsv.cell(r, C['L']).value, 'm': wsv.cell(r, C['M']).value,
            'n': wsv.cell(r, C['N']).value, 'o': wsv.cell(r, C['O']).value,
            'p': wsv.cell(r, C['P']).value, 'q': wsv.cell(r, C['Q']).value,
            'r': wsv.cell(r, C['R']).value, 'new': False,
        })
    return wb, ws, headers, C, U, existing


# ----------------------------------------------------------------------------- classify
def classify(new_rows, existing):
    # התאם מול כל שורת "אופציה" קיימת עם אותו גוש-חלקה מלא.
    # (היסטורית האופציות ניתנו ביוני 2025–מרץ 2026, אך ההתאמה לפי מזהה יחידה
    #  עמידה גם אם יינתנו אופציות בעתיד — אופציה תמיד מקדימה את המימוש.)
    optset = {e['a'] for e in existing if e['m'] == 'אופציה'}
    counts = {'מימוש אופציה': 0, 'דירה': 0}
    for row in new_rows:
        center = compound_of(row['a']) == 'מרכז'
        row['m'] = 'מימוש אופציה' if (center and row['a'] in optset) else 'דירה'
        counts[row['m']] += 1
    return counts, len(optset)


# ----------------------------------------------------------------------------- write back
def write_all(ws, headers, C, U, all_rows):
    # resolve every target column index — EXACT header match first, then startswith
    def ci(exact, starts=None):
        for i, h in enumerate(headers):
            if h is not None and str(h).strip() == exact:
                return i + 1
        if starts:
            for i, h in enumerate(headers):
                if h is not None and str(h).strip().startswith(starts):
                    return i + 1
        raise KeyError(exact)
    iA = ci('גוש חלקה'); iB = ci('גוש'); iC = ci('חלקה'); iD = ci('תת חלקה')
    iE = ci('מתחם'); iF = ci('יום מכירה'); iG = ci('יום'); iH = ci('חודש'); iI = ci('שנה')
    iJ = ci('תמורה מוצהרת בש"ח', starts='תמורה מוצהרת'); iK = ci('שווי מכירה בש"ח', starts='שווי מכירה')
    iL = ci('מהות'); iM = ci('סוג עסקה')
    iN = ci('חלק נמכר'); iO = ci('ישוב'); iP = ci('שנת בניה'); iQ = ci('שטח'); iR = ci('חדרים')
    iS = ci('מחיר למ״ר', starts='מחיר למ'); iT = ci('טווח תמורה מוצהרת', starts='טווח תמורה')

    old_max = ws.max_row
    r = 2
    for row in all_rows:
        ws.cell(r, iA).value = row['a']
        ws.cell(r, iB).value = f'=RIGHT(LEFT(A{r},6),4)'
        ws.cell(r, iC).value = f'=RIGHT(LEFT(A{r},11),3)'
        ws.cell(r, iD).value = f'=LEFT(RIGHT(A{r},6),3)'
        ws.cell(r, iE).value = compound_of(row['a'])
        cf = ws.cell(r, iF); cf.value = row['f']; cf.number_format = NUMFMT_DATE
        cg = ws.cell(r, iG); cg.value = f'=DAY(F{r})'; cg.number_format = '0'
        ws.cell(r, iH).value = f'=MONTH(F{r})'
        ws.cell(r, iI).value = f'=YEAR(F{r})'
        cj = ws.cell(r, iJ); cj.value = row['j']; cj.number_format = NUMFMT_ACCT0
        ck = ws.cell(r, iK); ck.value = row['k']; ck.number_format = NUMFMT_ACCT0
        ws.cell(r, iL).value = row['l']
        ws.cell(r, iM).value = row['m']
        ws.cell(r, iN).value = row['n']
        ws.cell(r, iO).value = row['o']
        ws.cell(r, iP).value = row['p']
        ws.cell(r, iQ).value = row['q']
        ws.cell(r, iR).value = row['r']
        cs = ws.cell(r, iS); cs.value = f'=J{r}/Q{r}'; cs.number_format = NUMFMT_ACCT2
        ws.cell(r, iT).value = (f'=IF(J{r}<$U$2,$U$2,IF(J{r}<$U$3,$U$3,'
                                f'IF(J{r}<$U$4,$U$4,$U$5)))')
        r += 1
    # clear any leftover old rows (cols A..T only — never touch U thresholds)
    last_written = r - 1
    for rr in range(last_written + 1, old_max + 1):
        for cc in range(1, iT + 1):
            ws.cell(rr, cc).value = None
    return last_written


def main():
    if len(sys.argv) < 3:
        print('usage: insert_into_master.py <master.xlsx> <merged_tax.xlsx> [output.xlsx]')
        sys.exit(1)
    master_path, merged_path = sys.argv[1], sys.argv[2]
    out_path = sys.argv[3] if len(sys.argv) > 3 else master_path

    log = []
    tax_rows = read_tax_rows(merged_path)
    log.append(f"שורות בקובץ המאוחד: {len(tax_rows)}")
    new_rows = clean_tax_rows(tax_rows, log)
    log.append(f"שורות חדשות אחרי ניקוי: {len(new_rows)}")

    wb, ws, headers, C, U, existing = read_master(master_path)
    log.append(f"שורות קיימות במאסטר: {len(existing)}")

    counts, nopt = classify(new_rows, existing)
    log.append(f"סיווג סוג עסקה (חדשות): מימוש אופציה {counts['מימוש אופציה']} · "
               f"דירה {counts['דירה']}  (התאמה מול {nopt} אופציות)")

    all_rows = existing + new_rows
    all_rows.sort(key=lambda x: (x['f'] or datetime.datetime(1900, 1, 1)))
    last = write_all(ws, headers, C, U, all_rows)
    log.append(f"סה\"כ שורות נתונים אחרי מיזוג ומיון: {last - 1}")

    wb.save(out_path)
    print('\n'.join(log))
    print(f"נשמר: {out_path}")


if __name__ == '__main__':
    main()
