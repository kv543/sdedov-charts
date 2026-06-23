#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
איחוד קבצי עסקאות נדל"ן שהורדו ממערכת מידע נדל"ן של רשות המיסים.

הקבצים שמורדים מהאתר נשמרים עם סיומת .xls אבל הם למעשה קובצי HTML
(טבלת HTML אחת לכל קובץ). הסקריפט קורא את כל הקבצים שבתיקייה, מאחד
את כל השורות לקובץ אחד, ממיין לפי תאריך, מסמן שורות שאינן ברובע שדה דב,
ושומר כ-xlsx.

שימוש:
    python3 merge_tax_files.py <input_dir> [output_xlsx]

אם output_xlsx לא צוין, נשמר כ"עסקאות מאוחדות <חודש שנה>.xlsx" בתוך input_dir.
"""
import sys
import os
import glob
import pandas as pd

# העמודות הצפויות בייצוא של רשות המיסים (בסדר הקבוע)
EXPECTED_COLS = [
    'גוש חלקה', 'יום מכירה', 'תמורה מוצהרת בש"ח', 'שווי מכירה בש"ח',
    'מהות', 'חלק נמכר', 'ישוב', 'שנת בניה', 'שטח', 'חדרים',
]
DATE_COL = 'יום מכירה'
GUSH_COL = 'גוש חלקה'
FLAG_COL = 'מחוץ לרובע'
FLAG_TEXT = '✓ למחיקה'
HEB_MONTHS = {1: 'ינואר', 2: 'פברואר', 3: 'מרץ', 4: 'אפריל', 5: 'מאי', 6: 'יוני',
              7: 'יולי', 8: 'אוגוסט', 9: 'ספטמבר', 10: 'אוקטובר', 11: 'נובמבר', 12: 'דצמבר'}

# ---------------------------------------------------------------------------
# רשימת החרגה: חלקות שאינן שייכות לרובע שדה דב (יש לסמן אותן למחיקה ידנית).
# מבנה: גוש -> אחד מהבאים:
#   * set של מספרי חלקות שמוחרגות
#   * הדיקט {'all_except': set} — כל החלקות בגוש מוחרגות פרט לאלו שברשימה
# גושים שאינם מופיעים כאן נחשבים שייכים לרובע במלואם.
# ---------------------------------------------------------------------------
EXCLUDE_CHELKOT = {
    6634: {333, 334, 335, 336, 348, 351},
    6896: {207, 208, 209, 210, 211, 212, 28, 29, 38, 158, 167, 12,
           5, 6, 7, 8, 9, 10, 11, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 2, 3},
    7186: {'all_except': {3}},
}


def parse_gush_chelka(val):
    """מפרק '006634-0015-013-00' ל-(גוש, חלקה) כמספרים שלמים. None אם לא ניתן."""
    try:
        parts = str(val).split('-')
        return int(parts[0]), int(parts[1])
    except Exception:
        return None, None


def is_outside_quarter(val):
    """מחזיר True אם השורה אינה שייכת לרובע שדה דב לפי EXCLUDE_CHELKOT."""
    gush, chelka = parse_gush_chelka(val)
    if gush is None:
        return False
    rule = EXCLUDE_CHELKOT.get(gush)
    if rule is None:
        return False
    if isinstance(rule, dict) and 'all_except' in rule:
        return chelka not in rule['all_except']
    return chelka in rule


def read_one(path):
    """קורא קובץ בודד. מנסה קודם כ-HTML (הפורמט של רשות המיסים), ואז כאקסל אמיתי."""
    try:
        tables = pd.read_html(path, encoding='utf-8')
        if tables:
            return tables[0]
    except Exception:
        pass
    return pd.read_excel(path)


def merge(input_dir, output_path=None):
    files = []
    for p in ('*.xls', '*.xlsx'):
        files += glob.glob(os.path.join(input_dir, p))
    files = sorted(f for f in files if 'מאוחד' not in os.path.basename(f))
    if not files:
        raise SystemExit(f'לא נמצאו קבצי .xls/.xlsx בתיקייה: {input_dir}')

    frames = []
    print(f'נמצאו {len(files)} קבצים:')
    for f in files:
        df = read_one(f)
        df.columns = [str(c).strip() for c in df.columns]
        missing = [c for c in EXPECTED_COLS if c not in df.columns]
        if missing:
            print(f'  ! אזהרה: בקובץ "{os.path.basename(f)}" חסרות עמודות: {missing}')
        df['_src'] = os.path.basename(f)
        frames.append(df)
        print(f'  • {os.path.basename(f)}: {len(df)} שורות')

    merged = pd.concat(frames, ignore_index=True)

    # מיון לפי תאריך (שמירה על המחרוזת המקורית בפלט)
    merged['_d'] = pd.to_datetime(merged[DATE_COL], format='%d/%m/%Y', errors='coerce')
    n_bad = int(merged['_d'].isna().sum())
    if n_bad:
        print(f'  ! אזהרה: {n_bad} שורות עם תאריך לא תקין (יופיעו בסוף)')
    merged = merged.sort_values('_d', kind='stable', na_position='last').reset_index(drop=True)

    # סימון שורות מחוץ לרובע
    merged[FLAG_COL] = merged[GUSH_COL].apply(lambda v: FLAG_TEXT if is_outside_quarter(v) else '')
    n_out = int((merged[FLAG_COL] == FLAG_TEXT).sum())

    # סיכום
    valid = merged['_d'].dropna()
    print(f'\nסה"כ שורות מאוחדות: {len(merged)}')
    if len(valid):
        print(f'טווח תאריכים: {valid.min().date()} עד {valid.max().date()}')
        by_month = valid.dt.to_period('M').value_counts().sort_index()
        print('פילוח לפי חודש:')
        for per, cnt in by_month.items():
            print(f'  {HEB_MONTHS.get(per.month, per.month)} {per.year}: {cnt} שורות')
    print(f'שורות מחוץ לרובע שדה דב (מסומנות למחיקה): {n_out}')
    dups = int(merged.duplicated(subset=EXPECTED_COLS).sum())
    if dups:
        print(f'הערה: {dups} שורות כפולות (זהות לחלוטין) — נשמרו כפי שהן.')

    # פלט: עמודות רשות המיסים + עמודת הסימון בסוף
    out_cols = [c for c in EXPECTED_COLS if c in merged.columns] + [FLAG_COL]
    out = merged[out_cols].copy()
    out[GUSH_COL] = out[GUSH_COL].astype(str)  # שמירה על אפסים מובילים כטקסט

    if output_path is None:
        if len(valid):
            mn, mx = valid.min(), valid.max()
            if (mn.year, mn.month) == (mx.year, mx.month):
                tag = f'{HEB_MONTHS[mn.month]} {mn.year}'
            else:
                tag = f'{HEB_MONTHS[mn.month]}-{HEB_MONTHS[mx.month]} {mx.year}'
            fname = f'עסקאות מאוחדות {tag}.xlsx'
        else:
            fname = 'עסקאות מאוחדות.xlsx'
        output_path = os.path.join(input_dir, fname)

    out.to_excel(output_path, index=False)
    _highlight_outside(output_path, out, FLAG_COL, FLAG_TEXT)
    print(f'\nנשמר: {output_path}')
    return output_path


def _highlight_outside(path, df, flag_col, flag_text):
    """צובע באדום בהיר את השורות המסומנות מחוץ לרובע, להבלטה ויזואלית."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    red = PatternFill(start_color='FFF4CCCC', end_color='FFF4CCCC', fill_type='solid')
    bold = Font(bold=True)
    # שורת כותרת מודגשת
    for c in ws[1]:
        c.font = bold
    flag_idx = list(df.columns).index(flag_col) + 1  # 1-based
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, flag_idx).value == flag_text:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = red
    # רוחב עמודות בסיסי
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(flag_idx)].width = 14
    wb.save(path)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        raise SystemExit('שימוש: python3 merge_tax_files.py <input_dir> [output_xlsx]')
    inp = sys.argv[1]
    outp = sys.argv[2] if len(sys.argv) > 2 else None
    merge(inp, outp)
